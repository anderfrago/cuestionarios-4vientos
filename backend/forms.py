from datetime import datetime, timezone
from html import escape
from io import BytesIO

from flask import Blueprint, abort, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .auth import current_user, roles_required
from .extensions import db
from .models import (Course, CourseQuestionnaire, CriticalAlert, Enrollment, FormAspect,
                     FormAttempt, FormResponse, Question, Questionnaire, Attempt,
                     QuestionnaireVersion, QuestionOption, QuestionRow)
from .routes import attempt_dict as legacy_attempt_dict, delete_form_attempts

forms = Blueprint("forms", __name__, url_prefix="/api")
QUESTION_TYPES = {"yes_no", "select", "text", "radio", "matrix", "number_matrix"}


def data():
    return request.get_json(silent=True) or {}


def fail(message, status=400):
    return jsonify(error=message), status


def can_view_course(user, course):
    return user.role == "admin" or (user.role == "tutor" and course.tutor_id == user.id)


def form_attempt_dict(attempt, include_responses=False):
    values = {}
    for response in attempt.responses:
        if response.score is not None and response.question.is_scored:
            values.setdefault(response.question.form_aspect, []).append(response.score)
    results = []
    for aspect, scores in values.items():
        average = round(sum(scores) / len(scores), 2)
        if average <= aspect.low_max:
            level, message = "Incipiente", aspect.low_message
        elif average <= aspect.medium_max:
            level, message = "En desarrollo", aspect.medium_message
        else:
            level, message = "Generado", aspect.high_message
        results.append({"aspect_id": aspect.id, "aspect": aspect.name, "average": average,
                        "level": level, "message": message})
    result = {"id": attempt.id, "created_at": attempt.created_at.isoformat(),
              "source": "versioned",
              "student": attempt.student.as_dict(), "course": attempt.course.as_dict(),
              "questionnaire": attempt.version.questionnaire.as_dict(),
              "version": attempt.version.version, "encouragement": attempt.encouragement,
              "results": sorted(results, key=lambda r: r["aspect_id"])}
    if include_responses:
        result["responses"] = [{"question_id": r.question_id, "question": r.question.title,
            "aspect": r.question.form_aspect.name, "row": r.row.label if r.row else None,
            "answer": r.option.label if r.option else r.text_value, "score": r.score}
            for r in attempt.responses]
    return result


def copy_version(source, questionnaire):
    version = QuestionnaireVersion(questionnaire_id=questionnaire.id,
        version=max([v.version for v in questionnaire.versions] or [0]) + 1, status="draft")
    db.session.add(version); db.session.flush()
    if not source:
        return version
    for old_aspect in source.aspects:
        if old_aspect.is_archived:
            continue
        aspect = FormAspect(version_id=version.id, name=old_aspect.name,
            description=old_aspect.description, order=old_aspect.order,
            low_max=old_aspect.low_max, medium_max=old_aspect.medium_max,
            low_message=old_aspect.low_message, medium_message=old_aspect.medium_message,
            high_message=old_aspect.high_message)
        db.session.add(aspect); db.session.flush()
        for old_question in old_aspect.questions:
            if old_question.is_archived:
                continue
            question = Question(aspect_id=aspect.id, title=old_question.title,
                help_text=old_question.help_text, question_type=old_question.question_type,
                required=old_question.required, order=old_question.order,
                reverse_scored=old_question.reverse_scored, is_scored=old_question.is_scored,
                allow_other=old_question.allow_other, is_critical=old_question.is_critical,
                critical_score_min=old_question.critical_score_min)
            db.session.add(question); db.session.flush()
            for option in old_question.options:
                if option.is_archived:
                    continue
                db.session.add(QuestionOption(question_id=question.id, label=option.label,
                    value=option.value, score=option.score, order=option.order))
            for row in old_question.rows:
                if row.is_archived:
                    continue
                db.session.add(QuestionRow(question_id=question.id, label=row.label, order=row.order))
    return version


@forms.route("/admin/questionnaires", methods=["GET", "POST"])
@roles_required("admin")
def questionnaires_admin():
    if request.method == "GET":
        return [q.as_dict(include_versions=True) for q in Questionnaire.query.order_by(Questionnaire.name)]
    body = data()
    questionnaire = Questionnaire(name=body.get("name", "").strip(),
        description=body.get("description", ""), level=int(body.get("level", 1)))
    if not questionnaire.name:
        return fail("El nombre es obligatorio")
    db.session.add(questionnaire); db.session.flush(); copy_version(None, questionnaire)
    db.session.commit(); return questionnaire.as_dict(include_versions=True), 201


@forms.route("/admin/questionnaires/<int:questionnaire_id>", methods=["GET", "PUT", "DELETE"])
@roles_required("admin")
def questionnaire_admin(questionnaire_id):
    questionnaire = Questionnaire.query.get_or_404(questionnaire_id)
    if request.method == "GET": return questionnaire.as_dict(include_versions=True)
    if request.method == "DELETE":
        if request.args.get("permanent", "false").lower() == "true":
            if not questionnaire.is_archived:
                return fail("Primero debes eliminar el formulario", 409)
            version_ids = [version.id for version in questionnaire.versions]
            if version_ids:
                delete_form_attempts(FormAttempt.query.filter(FormAttempt.version_id.in_(version_ids)))
            CourseQuestionnaire.query.filter_by(questionnaire_id=questionnaire.id).delete()
            db.session.delete(questionnaire)
            db.session.commit()
            return "", 204
        questionnaire.is_archived = True; db.session.commit(); return "", 204
    body = data()
    for key in ("name", "description", "level", "is_archived"):
        if key in body: setattr(questionnaire, key, body[key])
    db.session.commit(); return questionnaire.as_dict(include_versions=True)


@forms.post("/admin/questionnaires/<int:questionnaire_id>/versions")
@roles_required("admin")
def create_version(questionnaire_id):
    questionnaire = Questionnaire.query.get_or_404(questionnaire_id)
    source_id = data().get("source_version_id")
    source = db.session.get(QuestionnaireVersion, source_id) if source_id else questionnaire.published_version
    if source and source.questionnaire_id != questionnaire.id:
        return fail("La versión de origen no pertenece al formulario")
    version = copy_version(source, questionnaire); db.session.commit()
    return version.as_dict(), 201


@forms.post("/admin/versions/<int:version_id>/publish")
@roles_required("admin")
def publish_version(version_id):
    version = QuestionnaireVersion.query.get_or_404(version_id)
    if not any(a.questions for a in version.aspects): return fail("El formulario no contiene preguntas")
    for previous in version.questionnaire.versions:
        if previous.status == "published": previous.status = "superseded"
    version.status = "published"; version.published_at = datetime.now(timezone.utc)
    db.session.commit(); return version.as_dict()


def draft_version(version_id):
    version = QuestionnaireVersion.query.get_or_404(version_id)
    if version.status != "draft": abort(409, description="Las versiones publicadas son inmutables")
    return version


@forms.post("/admin/versions/<int:version_id>/aspects")
@roles_required("admin")
def create_aspect(version_id):
    version = draft_version(version_id); body = data()
    aspect = FormAspect(version_id=version.id, name=body.get("name", "Nuevo aspecto"),
                        description=body.get("description", ""), order=len(version.aspects) + 1)
    db.session.add(aspect); db.session.commit(); return aspect.as_dict(), 201


@forms.route("/admin/form-aspects/<int:aspect_id>", methods=["PUT", "DELETE"])
@roles_required("admin")
def edit_aspect(aspect_id):
    aspect = FormAspect.query.get_or_404(aspect_id); draft_version(aspect.version_id)
    if request.method == "DELETE": aspect.is_archived = True; db.session.commit(); return "", 204
    body = data()
    mapping = {"incipiente": "low_message", "en_desarrollo": "medium_message", "generado": "high_message"}
    for key in ("name", "description", "order", "low_max", "medium_max"):
        if key in body: setattr(aspect, key, body[key])
    for key, attr in mapping.items():
        if key in body.get("messages", {}): setattr(aspect, attr, body["messages"][key])
    db.session.commit(); return aspect.as_dict()


@forms.post("/admin/form-aspects/<int:aspect_id>/questions")
@roles_required("admin")
def create_question(aspect_id):
    aspect = FormAspect.query.get_or_404(aspect_id); draft_version(aspect.version_id); body = data()
    qtype = body.get("question_type", "radio")
    if qtype not in QUESTION_TYPES: return fail("Tipo de pregunta no válido")
    question = Question(aspect_id=aspect.id, title=body.get("title", "Nueva pregunta"),
        help_text=body.get("help_text", ""), question_type=qtype, required=body.get("required", True),
        order=len(aspect.questions) + 1, reverse_scored=body.get("reverse_scored", False),
        is_scored=body.get("is_scored", True), allow_other=body.get("allow_other", False),
        is_critical=body.get("is_critical", False), critical_score_min=body.get("critical_score_min"))
    db.session.add(question); db.session.commit(); return question.as_dict(), 201


@forms.route("/admin/questions/<int:question_id>", methods=["PUT", "DELETE"])
@roles_required("admin")
def edit_question(question_id):
    question = Question.query.get_or_404(question_id); draft_version(question.form_aspect.version_id)
    if request.method == "DELETE": question.is_archived = True; db.session.commit(); return "", 204
    body = data(); qtype = body.get("question_type", question.question_type)
    if qtype not in QUESTION_TYPES: return fail("Tipo de pregunta no válido")
    for key in ("title", "help_text", "question_type", "required", "order", "reverse_scored",
                "is_scored", "allow_other", "is_critical", "critical_score_min"):
        if key in body: setattr(question, key, body[key])
    if "options" in body:
        question.options.clear(); db.session.flush()
        for index, option in enumerate(body["options"], 1):
            db.session.add(QuestionOption(question_id=question.id, label=option["label"],
                value=str(option.get("value", index)), score=option.get("score"), order=index))
    if "rows" in body:
        question.rows.clear(); db.session.flush()
        for index, row in enumerate(body["rows"], 1):
            label = row["label"] if isinstance(row, dict) else row
            db.session.add(QuestionRow(question_id=question.id, label=label, order=index))
    db.session.commit(); return question.as_dict()


@forms.post("/admin/questions/<int:question_id>/move")
@roles_required("admin")
def move_question(question_id):
    question = Question.query.get_or_404(question_id); draft_version(question.form_aspect.version_id)
    direction = data().get("direction"); siblings = [q for q in question.form_aspect.questions if not q.is_archived]
    siblings.sort(key=lambda q: q.order); index = siblings.index(question); target = index + (-1 if direction == "up" else 1)
    if 0 <= target < len(siblings):
        question.order, siblings[target].order = siblings[target].order, question.order; db.session.commit()
    return question.as_dict()


@forms.put("/admin/courses/<int:course_id>/questionnaires")
@roles_required("admin")
def assign_questionnaires(course_id):
    course = Course.query.get_or_404(course_id); ids = {int(v) for v in data().get("questionnaire_ids", [])}
    valid = {q.id for q in Questionnaire.query.filter(Questionnaire.id.in_(ids),
        Questionnaire.level == course.level, Questionnaire.is_archived.is_(False)).all()} if ids else set()
    if ids != valid: return fail("Algún formulario no existe o no corresponde al nivel del curso")
    CourseQuestionnaire.query.filter_by(course_id=course.id).delete()
    db.session.add_all([CourseQuestionnaire(course_id=course.id, questionnaire_id=i) for i in valid])
    db.session.commit(); return {"questionnaire_ids": sorted(valid)}


@forms.get("/courses/<int:course_id>/forms")
@roles_required("student", "tutor", "admin")
def course_forms(course_id):
    course, user = Course.query.get_or_404(course_id), current_user()
    if user.role == "student" and not Enrollment.query.filter_by(student_id=user.id, course_id=course.id).first(): abort(403)
    if user.role != "student" and not can_view_course(user, course): abort(403)
    assigned = CourseQuestionnaire.query.filter_by(course_id=course.id, is_active=True).all()
    forms_data = []
    for assignment in assigned:
        q = assignment.questionnaire; version = q.published_version
        if not version or q.is_archived: continue
        count = FormAttempt.query.filter_by(student_id=user.id, course_id=course.id, version_id=version.id).count() if user.role == "student" else 0
        item = q.as_dict(); item.update({"version_id": version.id, "version": version.version, "attempt_count": count})
        forms_data.append(item)
    return {"course": course.as_dict(), "forms": forms_data}


@forms.get("/courses/<int:course_id>/forms/<int:version_id>")
@roles_required("student", "tutor", "admin")
def form_definition(course_id, version_id):
    course, user = Course.query.get_or_404(course_id), current_user()
    if user.role == "student" and not Enrollment.query.filter_by(student_id=user.id, course_id=course.id).first(): abort(403)
    if user.role != "student" and not can_view_course(user, course): abort(403)
    version = QuestionnaireVersion.query.get_or_404(version_id)
    assigned = CourseQuestionnaire.query.filter_by(course_id=course.id,
        questionnaire_id=version.questionnaire_id, is_active=True).first()
    if not assigned or version.status != "published": abort(403)
    return {"course": course.as_dict(), "questionnaire": version.questionnaire.as_dict(), "version": version.as_dict()}


@forms.post("/courses/<int:course_id>/forms/<int:version_id>/attempts")
@roles_required("student")
def submit_form(course_id, version_id):
    course, user = Course.query.get_or_404(course_id), current_user()
    if not Enrollment.query.filter_by(student_id=user.id, course_id=course.id).first(): abort(403)
    version = QuestionnaireVersion.query.get_or_404(version_id)
    if version.status != "published" or not CourseQuestionnaire.query.filter_by(course_id=course.id,
            questionnaire_id=version.questionnaire_id, is_active=True).first(): abort(403)
    submitted = data().get("responses", []); by_key = {(int(r["question_id"]), r.get("row_id")): r for r in submitted}
    questions = [q for a in version.aspects if not a.is_archived for q in a.questions if not q.is_archived]
    for question in questions:
        keys = [(question.id, row.id) for row in question.rows if not row.is_archived] or [(question.id, None)]
        if question.required and any(key not in by_key or not (by_key[key].get("option_id") or
                str(by_key[key].get("text_value", "")).strip()) for key in keys):
            return fail(f"Falta responder: {question.title}")
    attempt = FormAttempt(student_id=user.id, course_id=course.id, version_id=version.id,
        encouragement="Gracias por escucharte. Reconocer cómo estás es el primer paso para seguir creciendo.")
    db.session.add(attempt); db.session.flush(); alerts = []
    for raw in submitted:
        question = db.session.get(Question, int(raw["question_id"]))
        if not question or question.form_aspect.version_id != version.id: return fail("Respuesta no válida")
        row_id = raw.get("row_id"); option_id = raw.get("option_id"); option = db.session.get(QuestionOption, option_id) if option_id else None
        row = db.session.get(QuestionRow, row_id) if row_id else None
        if row and row.question_id != question.id: return fail("Fila no válida")
        if option and option.question_id != question.id: return fail("Opción no válida")
        score = option.score if option else None
        if score is not None and question.reverse_scored:
            scored = [o.score for o in question.options if o.score is not None]
            score = min(scored) + max(scored) - score if scored else score
        response = FormResponse(attempt_id=attempt.id, question_id=question.id, row_id=row_id,
            option_id=option_id, text_value=str(raw.get("text_value", "")).strip() or None,
            score=score if question.is_scored else None)
        db.session.add(response); db.session.flush()
        if question.is_critical and score is not None and score >= (question.critical_score_min or 1):
            alerts.append(CriticalAlert(attempt_id=attempt.id, response_id=response.id))
    db.session.add_all(alerts); db.session.commit(); return form_attempt_dict(attempt, True), 201


@forms.get("/courses/<int:course_id>/form-analytics")
@roles_required("tutor", "admin")
def form_analytics(course_id):
    course, user = Course.query.get_or_404(course_id), current_user()
    if not can_view_course(user, course): abort(403)
    attempts = FormAttempt.query.filter_by(course_id=course.id).order_by(FormAttempt.created_at).all()
    detail = [form_attempt_dict(a, True) for a in attempts]
    legacy = Attempt.query.filter_by(course_id=course.id).order_by(Attempt.created_at).all()
    for attempt in legacy:
        row = legacy_attempt_dict(attempt, True); row["source"] = "legacy"; detail.append(row)
    detail.sort(key=lambda row: row["created_at"])
    totals = {}
    for attempt in detail:
        for result in attempt["results"]: totals.setdefault(result["aspect"], []).append(result["average"])
    summary = [{"aspect": k, "average": round(sum(v)/len(v), 2), "count": len(v)} for k, v in totals.items()]
    alerts = CriticalAlert.query.join(FormAttempt).filter(FormAttempt.course_id == course.id).order_by(CriticalAlert.created_at.desc()).all()
    return {"course": course.as_dict(), "summary": summary, "attempts": detail,
            "alerts": [a.as_dict() for a in alerts]}


@forms.route("/alerts/<int:alert_id>/review", methods=["PUT"])
@roles_required("tutor", "admin")
def review_alert(alert_id):
    alert, user = CriticalAlert.query.get_or_404(alert_id), current_user()
    if not can_view_course(user, alert.attempt.course): abort(403)
    alert.reviewed_at = datetime.now(timezone.utc); alert.reviewed_by_id = user.id
    alert.review_notes = data().get("notes", ""); db.session.commit(); return alert.as_dict()


def export_attempts(course):
    return FormAttempt.query.filter_by(course_id=course.id).order_by(FormAttempt.created_at).all()


@forms.get("/courses/<int:course_id>/export.xlsx")
@roles_required("tutor", "admin")
def export_xlsx(course_id):
    course, user = Course.query.get_or_404(course_id), current_user()
    if not can_view_course(user, course): abort(403)
    wb = Workbook(); ws = wb.active; ws.title = "Respuestas"
    headers = ["Formulario", "Versión", "Alumno/a", "Correo", "Fecha", "Aspecto", "Pregunta", "Fila", "Respuesta", "Puntuación"]
    ws.append(headers)
    for cell in ws[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="0B5D78")
    for attempt in export_attempts(course):
        for response in attempt.responses:
            ws.append([attempt.version.questionnaire.name, attempt.version.version, attempt.student.name,
                attempt.student.email, attempt.created_at.replace(tzinfo=None), response.question.form_aspect.name,
                response.question.title, response.row.label if response.row else "",
                response.option.label if response.option else response.text_value or "", response.score])
    for attempt in Attempt.query.filter_by(course_id=course.id).order_by(Attempt.created_at):
        for answer in attempt.answers:
            score = 5 - answer.value if answer.item.reverse_scored else answer.value
            ws.append(["Cuestionario histórico", "Anterior", attempt.student.name,
                attempt.student.email, attempt.created_at.replace(tzinfo=None), answer.item.aspect.name,
                answer.item.text, "", answer.value, score])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"; ws.sheet_view.showGridLines = False
    widths = [24,10,22,30,19,24,55,34,45,12]
    for index, width in enumerate(widths, 1): ws.column_dimensions[chr(64+index)].width = width
    for row in ws.iter_rows():
        for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
    out = BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=f"{course.name}-respuestas.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def pdf_document(course, attempts, title, legacy_attempts=None):
    out = BytesIO(); doc = SimpleDocTemplate(out, pagesize=landscape(A4), rightMargin=1.2*cm,
        leftMargin=1.2*cm, topMargin=1.2*cm, bottomMargin=1.2*cm)
    styles = getSampleStyleSheet(); story = [Paragraph(escape(title), styles["Title"]),
        Paragraph(escape(f"{course.name} · {course.academic_year}"), styles["Heading2"]), Spacer(1, 10)]
    aggregates = {}
    for attempt in attempts:
        for result in form_attempt_dict(attempt)["results"]:
            aggregates.setdefault(result["aspect"], []).append(result["average"])
    if aggregates:
        summary_rows = [["Aspecto", "Media", "Mediciones"]] + [[name,
            f"{sum(values)/len(values):.2f}", str(len(values))] for name, values in aggregates.items()]
        summary = Table(summary_rows, colWidths=[10*cm, 3*cm, 3*cm], repeatRows=1)
        summary.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0B5D78")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#CCD7E0")),
            ("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),5)]))
        story.extend([Paragraph("Resumen estadístico", styles["Heading2"]), summary, Spacer(1, 14)])
    for attempt in attempts:
        story.extend([Paragraph(escape(f"{attempt.student.name} · {attempt.version.questionnaire.name} v{attempt.version.version}"), styles["Heading2"]),
                      Paragraph(attempt.created_at.strftime("%d/%m/%Y %H:%M"), styles["Normal"])])
        rows = [["Aspecto", "Pregunta", "Fila", "Respuesta", "Puntuación"]]
        for response in attempt.responses:
            rows.append([response.question.form_aspect.name, response.question.title,
                response.row.label if response.row else "", response.option.label if response.option else response.text_value or "",
                "" if response.score is None else f"{response.score:g}"])
        table = Table([[Paragraph(escape(str(c)), styles["BodyText"]) for c in row] for row in rows],
                      colWidths=[4*cm,7*cm,5*cm,7*cm,2*cm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0B5D78")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#CCD7E0")),
            ("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),
            ("TOPPADDING",(0,0),(-1,-1),4),("BOTTOMPADDING",(0,0),(-1,-1),4)]))
        story.extend([table, Spacer(1, 14)])
    for attempt in legacy_attempts or []:
        story.extend([Paragraph(escape(f"{attempt.student.name} · Cuestionario histórico"), styles["Heading2"]),
                      Paragraph(attempt.created_at.strftime("%d/%m/%Y %H:%M"), styles["Normal"])])
        rows = [["Aspecto", "Pregunta", "Respuesta", "Puntuación"]]
        for answer in attempt.answers:
            score = 5 - answer.value if answer.item.reverse_scored else answer.value
            rows.append([answer.item.aspect.name, answer.item.text, str(answer.value), str(score)])
        table = Table([[Paragraph(escape(str(c)), styles["BodyText"]) for c in row] for row in rows],
                      colWidths=[5*cm,12*cm,4*cm,3*cm], repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#0B5D78")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#CCD7E0")),
            ("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),5)]))
        story.extend([table, Spacer(1, 14)])
    doc.build(story); out.seek(0); return out


@forms.get("/courses/<int:course_id>/export.pdf")
@roles_required("tutor", "admin")
def export_pdf(course_id):
    course, user = Course.query.get_or_404(course_id), current_user()
    if not can_view_course(user, course): abort(403)
    legacy = Attempt.query.filter_by(course_id=course.id).order_by(Attempt.created_at).all()
    out = pdf_document(course, export_attempts(course), "Informe de autopercepción del curso", legacy)
    return send_file(out, as_attachment=True, download_name=f"{course.name}-informe.pdf", mimetype="application/pdf")


@forms.get("/attempts/<int:attempt_id>/export.pdf")
@roles_required("tutor", "admin")
def export_attempt_pdf(attempt_id):
    attempt, user = FormAttempt.query.get_or_404(attempt_id), current_user()
    if not can_view_course(user, attempt.course): abort(403)
    out = pdf_document(attempt.course, [attempt], "Ficha individual de autopercepción")
    return send_file(out, as_attachment=True, download_name=f"intento-{attempt.id}.pdf", mimetype="application/pdf")
